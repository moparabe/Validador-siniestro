"""
file_service.py
===============
Servicio de manejo de archivos:
  - Lectura segura del Excel cargado
  - Generación del archivo de salida limpio para Power BI
  - Generación del reporte de errores para devolver al ajustador
  - Subida a GitHub con control de versiones y sha

Separado de la UI para facilitar pruebas y reutilización.
"""

from __future__ import annotations

import base64
import io
from datetime import datetime
from typing import Optional

import pandas as pd
import requests

from validation_config import COLUMNAS_REQUERIDAS, COLUMNAS_SISTEMA


# ---------------------------------------------------------------------------
# LECTURA SEGURA DEL ARCHIVO
# ---------------------------------------------------------------------------
def leer_excel(archivo_bytes: bytes, max_filas: int = 5000) -> tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    Lee el archivo Excel cargado de forma segura.

    Retorna (DataFrame, None) si tuvo éxito,
    o (None, mensaje_error) si falló.
    """
    try:
        df = pd.read_excel(io.BytesIO(archivo_bytes), engine="openpyxl")
    except Exception as e:
        return None, (
            f"No se pudo leer el archivo Excel. Asegúrese de que no esté "
            f"protegido con contraseña y sea un .xlsx válido. Detalle: {e}"
        )

    if df.empty:
        return None, "El archivo está vacío. Por favor cargue un archivo con datos."

    if len(df) > max_filas:
        return None, (
            f"El archivo tiene {len(df):,} filas, pero el máximo permitido es "
            f"{max_filas:,}. Divida el archivo y cárguelo en partes."
        )

    return df, None


# ---------------------------------------------------------------------------
# GENERACIÓN DE ARCHIVOS DE SALIDA
# ---------------------------------------------------------------------------

def generar_excel_validos(df_validos: pd.DataFrame, nombre_ajustador: str) -> bytes:
    """
    Genera el Excel limpio con solo los registros válidos.
    Este es el archivo que va al flujo de Power BI.

    Columnas de estado (_ESTADO_FILA, _ERRORES_FILA) se eliminan de la
    salida final para mantener la estructura estable del modelo.
    """
    columnas_salida = COLUMNAS_SISTEMA + COLUMNAS_REQUERIDAS + ["_TIMESTAMP_CARGA", "_MES_REPORTE"]
    # Incluir solo las que existen en el DataFrame
    cols_presentes = [c for c in columnas_salida if c in df_validos.columns]

    df_salida = df_validos[cols_presentes].copy()

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_salida.to_excel(writer, index=False, sheet_name="Siniestros_Validos")
        _aplicar_estilo_excel(writer, "Siniestros_Validos", df_salida)
    return buffer.getvalue()


def generar_excel_reporte_errores(
    df_rechazados: pd.DataFrame,
    lista_errores: list[dict],
    nombre_ajustador: str,
) -> bytes:
    """
    Genera el Excel de reporte de errores para devolver al ajustador.
    Contiene:
      - Hoja 'Registros_con_Errores': filas rechazadas con la columna de errores
      - Hoja 'Detalle_Errores': tabla fila/columna/mensaje para fácil corrección
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Hoja 1: registros rechazados con descripción del error
        if not df_rechazados.empty:
            cols_rechazo = (
                ["_ESTADO_FILA", "_ERRORES_FILA"]
                + [c for c in COLUMNAS_REQUERIDAS if c in df_rechazados.columns]
            )
            cols_rechazo = [c for c in cols_rechazo if c in df_rechazados.columns]
            df_rechazo_hoja = df_rechazados[cols_rechazo].copy()
            df_rechazo_hoja.to_excel(writer, index=True, index_label="Fila_Original",
                                     sheet_name="Registros_con_Errores")
            _aplicar_estilo_excel(writer, "Registros_con_Errores", df_rechazo_hoja)

        # Hoja 2: detalle de errores por columna
        if lista_errores:
            df_errores = pd.DataFrame(lista_errores)
            df_errores.columns = ["Fila_Excel", "Columna", "Descripción_del_Error", "Severidad"]
            df_errores["Cómo_Corregirlo"] = df_errores["Columna"].map(
                _generar_sugerencia_correccion
            )
            df_errores.to_excel(writer, index=False, sheet_name="Detalle_Errores")
            _aplicar_estilo_excel(writer, "Detalle_Errores", df_errores)

    return buffer.getvalue()


def generar_excel_completo(df_validado: pd.DataFrame) -> bytes:
    """
    Genera el Excel completo con todas las filas y su estado (VÁLIDO/RECHAZADO).
    Útil para auditoría y trazabilidad.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_validado.to_excel(writer, index=False, sheet_name="Todos_los_Registros")
        _aplicar_estilo_excel(writer, "Todos_los_Registros", df_validado)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# SUBIDA A GITHUB
# ---------------------------------------------------------------------------

def subir_a_github(
    nombre_archivo: str,
    contenido_bytes: bytes,
    token: str,
    owner: str,
    repo: str,
    branch: str = "main",
    folder: str = "cargas",
) -> tuple[bool, str]:
    """
    Sube un archivo a GitHub. Si ya existe, lo actualiza (versionamiento por sha).

    Retorna (éxito: bool, mensaje: str).
    """
    folder = folder.strip("/")
    path = f"{folder}/{nombre_archivo}"
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    # Verificar si el archivo ya existe para obtener sha
    r_get = requests.get(url, headers=headers, params={"ref": branch}, timeout=15)
    sha = r_get.json().get("sha") if r_get.status_code == 200 else None

    payload: dict = {
        "message": f"Carga desde Streamlit: {nombre_archivo}",
        "content": base64.b64encode(contenido_bytes).decode("utf-8"),
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha

    try:
        r_put = requests.put(url, headers=headers, json=payload, timeout=30)
    except requests.exceptions.Timeout:
        return False, "La conexión con GitHub tardó demasiado. Intente nuevamente."
    except requests.exceptions.ConnectionError:
        return False, "No se pudo conectar con GitHub. Verifique la conexión a internet."

    if r_put.status_code in (200, 201):
        html_url = r_put.json().get("content", {}).get("html_url", "")
        return True, html_url

    return False, (
        f"Error al subir a GitHub (código {r_put.status_code}): {r_put.text[:300]}"
    )


def construir_nombre_archivo(
    nombre_original: str,
    username: str,
    tipo: str = "VALIDOS",
) -> str:
    """
    Construye el nombre de archivo con timestamp y usuario para trazabilidad.
    Formato: YYYYMMDD_HHMMSS_USUARIO_TIPO_nombre_original.xlsx
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_limpio = nombre_original.replace(" ", "_").replace("/", "-")
    return f"{timestamp}_{username.upper()}_{tipo}_{nombre_limpio}"


# ---------------------------------------------------------------------------
# HELPERS PRIVADOS
# ---------------------------------------------------------------------------

def _aplicar_estilo_excel(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Aplica formato básico al Excel de salida:
    encabezado con fondo, autoajuste de columnas.
    """
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet_name]

        # Estilo de encabezado
        fill_header = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Autoajuste de ancho
        for i, col in enumerate(df.columns, 1):
            max_width = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_width + 4, 50)

        # Filas con error en rojo (si existe columna _ESTADO_FILA)
        if "_ESTADO_FILA" in df.columns:
            from openpyxl.styles import PatternFill as PF
            fill_error = PF(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            estado_col_idx = list(df.columns).index("_ESTADO_FILA") + 1
            for row in ws.iter_rows(min_row=2):
                estado_cell = row[estado_col_idx - 1]
                if estado_cell.value == "RECHAZADO":
                    for cell in row:
                        cell.fill = fill_error

    except Exception:
        pass  # El formato es cosmético; si falla, no interrumpir el flujo


def _generar_sugerencia_correccion(columna: str) -> str:
    """Devuelve una sugerencia de corrección amigable por columna."""
    sugerencias = {
        "SINIESTRO": "Ingrese exactamente 8 dígitos (ej: 12345678) o el formato SAN##### (ej: SAN12345).",
        "POLIZA": "Ingrese exactamente 9 dígitos sin letras ni espacios (ej: 123456789).",
        "RAMO": "Ingrese exactamente 3 dígitos iniciando con 0 (ej: 012, 034, 099).",
        "NIT/CC": "Ingrese solo dígitos, sin puntos ni guiones (ej: 9001234567).",
        "AJUSTE": "Use únicamente los valores: AUTOS o VIDA (en mayúsculas).",
        "ESTADO_DOCUMENTO": "Use únicamente los valores: RECIBIDO o PENDIENTE (en mayúsculas).",
        "RESERVA": "Ingrese un número positivo sin símbolos de moneda (ej: 1500000 o 1500000.50).",
    }
    if "FECHA" in columna:
        return "Use el formato dd/mm/yyyy (ej: 31/12/2024). No deje este campo vacío si es obligatorio."
    return sugerencias.get(columna, "Revise el valor ingresado según las instrucciones de la plantilla.")
